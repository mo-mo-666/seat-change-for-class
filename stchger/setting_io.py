import enum
import openpyxl
import os

from openpyxl.utils.cell import col


def read_setting(path: str, is_hopes: bool = True, read_glasses: bool = True) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True)
    desks_map_s = wb["desks_map"]
    members_s = wb["members"]
    datas = {}

    # read seat places
    name_coord = {}
    coord_name = {}
    for i, row in enumerate(desks_map_s.iter_rows(values_only=True), start=1):
        for j, cell in enumerate(row, start=1):
            if cell:
                v = str(cell).strip()
                if v:
                    name_coord[v] = (j, i)
                    coord_name[(j, i)] = v
    datas["desks_map"] = (name_coord, coord_name)

    # read glasses_desks
    if read_glasses:
        glasses_desks = []
        glasses_desks_s = wb["glasses_desks"]
        for row in glasses_desks_s.iter_rows(values_only=True):
            for cell in row:
                v = str(cell).strip()
                if v and v in name_coord:
                    glasses_desks.append(name_coord[v])
        datas["glasses_desks"] = glasses_desks

    # read members sheet
    key = []
    for cell in members_s["2"]:
        v = str(cell.value).strip()
        if v:
            key.append(v)
    datas["member_keys"] = key

    members_d = []
    for row in members_s.iter_rows(min_row=3, values_only=True):
        member = {}
        for i, (k, v) in enumerate(zip(key, row)):
            v = str(v).strip()
            if i == 0 and not v:
                break
            member[k] = v
        member["number"] = int(member["number"])
        members_d.append(member)

    # reformat member data

    if is_hopes:
        members_rd = []
        for member in members_d:
            hopes = member["hopes"].split(",")
            hopes = tuple([name_coord[h] for h in hopes])
            member["hopes"] = hopes
            if read_glasses:
                if member["glasses"] == "0":
                    member["glasses"] = ()
                else:
                    member["glasses"] = glasses_desks
            members_rd.append(member)
        datas["members"] = members_rd
    else:
        datas["members"] = members_d

    datas["seat_places"] = tuple(coord_name.keys())
    return datas


def write_setting(path: str, datas: dict):
    exist = os.path.exists(path)
    if exist:
        wb = openpyxl.load_workbook(path, read_only=False)
    else:
        wb = openpyxl.Workbook()

    members = datas["members"]
    name_coord, coord_name = datas["desks_map"]
    member_keys = datas["member_keys"]
    mem_places = datas["mem_places"]
    loss_log = datas["loss_log"]

    # write member list
    ws = wb.create_sheet("result_member")
    for i, k in enumerate(member_keys, start=1):
        ws.cell(row=2, column=i, value=k)
    for i, k in enumerate(member_keys, start=1):
        for j, m in enumerate(members, start=3):
            if k == "hopes":
                v = [coord_name[mm] for mm in m[k]]
                v = ",".join(v)
            elif k == "glasses":
                v = 1 if m[k] else 0
            elif k == "mem_place":
                v = coord_name[m[k]]
            else:
                v = m[k]
            ws.cell(row=j, column=i, value=v)

    # write number place map
    ws = wb.create_sheet("result_num_map")
    for i, (x, y) in enumerate(mem_places, start=1):
        ws.cell(row=y, column=x, value=i)

    # write name place map
    ws = wb.create_sheet("result_member_map")
    for m, (x, y) in zip(members, mem_places):
        ws.cell(row=y, column=x, value=m["name"])

    # write loss
    ws = wb.create_sheet("result_loss")
    ws.cell(row=1, column=1, value=loss_log[-1])

    wb.save(path)
